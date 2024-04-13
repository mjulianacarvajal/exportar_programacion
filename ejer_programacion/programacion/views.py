from django.shortcuts import render
from django.http.response import HttpResponse

from .models import *
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from openpyxl import Workbook
from openpyxl.styles import *





# Create your views here.


def is_valid_queryparam(param):
    return param != '' and param is not None  


def programacion_list(request):
    qs = Programacion.objects.order_by('programacion')
    programacion = request.GET.get(' programacion')
    codigo = request.GET.get('codigo')
    estado = request.GET.get('estado')
    origen = request.GET.get('origen')
    destino = request.GET.get('destino')
    
    if is_valid_queryparam(estado):
        qs = qs.filter(estado__icontains=estado)

    if is_valid_queryparam(codigo):
        qs = qs.filter(codigo=codigo)
        
    page = request.GET.get('page', 1)
    paginator = Paginator(qs, 30)

    try:
        qs = paginator.page(page)
    except PageNotAnInteger:
        qs = paginator.page(1)
    except EmptyPage:
        qs = paginator.page(paginator.num_pages)


    context = {
        'programacion_list': qs,
        'programacion': programacion,
        'codigo':codigo,
        'estado':estado,
        'origen':origen,
        'destino':destino,
        
    }    

    return render(request, "excelexport/programacion_list.html", context)


def programacion_excel(request):
    
    qs = Programacion.objects.order_by('programacion')


    if 'id' in request.session:
        id = request.session['id']
    else:
        id = None   

    if 'programacion' in request.session:
        programacion = request.session['programacion']
    else:
        programacion = None

    if 'codigo'  in request.session:
        codigo =  request.session['codigo'] 
    else:
        codigo = None         

    if 'precio'  in request.session:
        precio =  request.session['precio'] 
    else:
        precio = None    

    if 'estado'  in request.session:
        estado =  request.session['estado'] 
    else:
        estado = None  

        
    if is_valid_queryparam(codigo):
        qs = qs.filter(codigo__icontains=codigo)

    if is_valid_queryparam(estado):
       qs = qs.filter(estado=estado)     

    if estado is None or estado == '':
        estado = "Todas las Programaciones"
    else:
       estado = estado

    if estado is None or estado == '':
        estado = "All Countries"
    else:
        estado = estado

    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + 'Listado Programaciones' +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:G1')

    first_cell = worksheet['A1']
     

    first_cell.value = "Listado Programaciones"
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font  = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'Listado Programaciones' 

    # Titulos Columnas
    columns = ['id','programacion','codigo','estado','origen', 'destino', 'precio',]
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font  = Font(bold=True, color="F7F6FA")
        seventh_cell = worksheet['G3']
        seventh_cell.alignment = Alignment(horizontal="right")

    for programaciones in qs:
        row_num += 1

        # Define the data for each cell in the row
        
        row = [programaciones.id, programaciones.programacion.strftime("%d%m%Y - %H:%M:%S"), programaciones.codigo, programaciones.estado, programaciones.origen.sede, programaciones.destino.sede, programaciones.precio]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
                         

    workbook    

    workbook.save(response)
    return response